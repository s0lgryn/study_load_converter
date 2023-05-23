import pandas as pd
import os
import re
from typing import List, Any, Union

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from requirements.dependencies import COLUMNS


# TODO написать docstring к каждой функции
def main():
    files = os.listdir("study_plans/test_plans")
    for file in files:
        # file.replace(".plx", "")
        wb = openpyxl.load_workbook("study_plans/test_plans/" + file)
        title_sheet = wb["Титул"]
        plan_sheet = wb["План"]
        disciplines_sheet = wb["Компетенции"]
        entry_year = find_entry_year(title_sheet)
        print(entry_year)
        min_col, max_col = find_course_boundaries(sheet=plan_sheet, today_year=2024, entry_year=int(entry_year))
        disciplines_col = find_disciplines_column(sheet=plan_sheet)
        study_load = parse_study_load(sheet=plan_sheet, min_course_col=min_col, max_course_col=max_col,
                                      disciplines_col=disciplines_col)
        print(study_load)
        disciplines = (parse_disciplines(disciplines_sheet))
        print(disciplines)
        # unique_disciplines = disciplines["Шифр дисциплины"].unique()
        # print(unique_disciplines)
        # study_load_filtered = study_load[study_load["Шифр дисциплины"].isin(unique_disciplines)]
        # print(study_load_filtered.to_string())


def check_filenames(files: Any) -> list:
    """Проверяет список файлов с которыми мы будем работать на корректность названия

    :param files: Список файлов которые выбрал пользователь
    :type files: List
    :returns: Список файлов прошедших валидацию
    :rtype: list
    """
    valid_files = []
    invalid_files = []
    for file in files:
        if filename_validator(file):
            valid_files.append(file)
            print(f"Имя файла: {file} корректно")
        else:
            invalid_files.append(file)
            print(f"Имя файла: {file} не корректно")
    return valid_files


def filename_validator(filename: str) -> bool:
    """
    # Код.специальности_XXX_стандарт_количество.курсов_XXX_класс_год.набора_специальность
    # regex: d{2}.d{2}.d{2}_d{2}_d{2}_(d{3,4}/d{4})-2843_(d{2})-(d{4})

    :param filename: Имя файла которое будем валидировать
    :type filename: string
    :returns: Значение результата проверки
    :rtype: bool
    """
    if re.search(r"\d{2}\.\d{2}\.\d{2}[_-]\d{2}[_-]\d{2}[_-]\d{2,4}[_-]2843[_-]\d{2}[_-]\d{4}[_-][а-яА-Я]{1,3}",
                 filename):
        return True
    return False


def extract_data_from_filename(filename: str) -> dict[str, int]:
    """
    :param filename:
    :return:
    """

    namings = ["specialization", "fgos_standard", "course_numbers", "class_base", "entry_year"]
    name = filename.split("/")[-1]
    file_data = name.replace("_", "-").rstrip(".").split("-")[:-1]
    file_data.pop(1)  # Удаление ненужной части
    file_data.remove("2843")  # Удаления кода организации "2843"
    if len(file_data) == 5:
        file_dict = dict(zip(namings, file_data))
        return file_dict


def find_entry_year(title_sheet: Worksheet) -> int:
    pattern = re.compile(r"^\d{4}$")
    for row in title_sheet.iter_rows(min_row=20, max_row=60, values_only=True):
        year_cell = next((cell for cell in row if cell and re.match(pattern, str(cell))), None)
        if year_cell:
            return year_cell
    print("Не был найден год поступления.")


# TODO доделать что бы вывод возвращал 4 границы, по 2 для каждого семестра написать отдельную функцию поиска семестра
def find_course_boundaries(sheet: Worksheet, today_year: int, entry_year: int) -> Union[List[int], None]:
    cell: Cell
    course_number = today_year - entry_year + 1
    pattern = r"[а-яА-Я]{4}\s+" + str(course_number)
    try:
        for row in sheet.iter_rows(max_row=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and re.search(pattern, cell.value):
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            return [merged_range.min_col, merged_range.max_col]
    except Exception as e:
        print(f"Границы курса были не найдены. {e}")
    return None


def find_disciplines_column(sheet: Worksheet) -> int:
    cell: Cell
    pattern = r"\Наименование$"
    try:
        for row in sheet.iter_rows(max_row=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and re.search(pattern, cell.value):
                    return cell.column
    except Exception as e:
        print(f"Не найден столбец в котором содержатся дисциплины {e}")


# TODO может не захватывать данные типа: "ПДП", предусмотреть.
def parse_disciplines(sheet: Worksheet) -> DataFrame:
    cell: Cell
    columns = ["Шифр дисциплины", "Наименование дисциплины"]
    data = pd.DataFrame(columns=columns)
    pattern = r"[а-яА-Я]{1,}\.\d{1,3}"  # https://regex101.com/r/Htwvzy/1

    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value and isinstance(cell.value, str) and re.search(pattern, cell.value):
                data.loc[len(data)] = [cell.value, ""]
                # получаем индекс последней добавленной строки в DataFrame
                last_idx = len(data) - 1
                # добавляем значение следующей за текущей ячейки во второй столбец
                if i + 1 < len(row):
                    data.at[last_idx, "Наименование дисциплины"] = row[i + 1].value
    return data


# TODO понять как добавлять данные только из тех строк, которые хранятся в DF[Наименование дисциплины] из пред. функции
def parse_study_load(sheet: Worksheet, min_course_col: int, max_course_col: int, disciplines_col: int) -> DataFrame:
    cell: Cell
    columns = ["Шифр дисциплины", "Наименование дисциплины"]
    data = pd.DataFrame(columns=columns)
    pattern = r"[а-яА-Я]{1,}\.\d{1,3}"
    row_in_dataframe = 0
    for row in sheet.iter_rows(min_col=disciplines_col - 1, max_col=disciplines_col, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and re.search(pattern, cell):
                data.at[row_in_dataframe, "Шифр дисциплины"] = row[0]
                data.at[row_in_dataframe, "Наименование дисциплины"] = row[1]
                row_in_dataframe += 1
    return data


if __name__ == '__main__':
    main()
