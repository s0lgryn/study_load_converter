import pandas as pd
import re
from typing import List, Any, Optional
import openpyxl as opx
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from parsing.preparation import get_filepath
from requirements.dependencies import COLUMNS, SPECIALIZATIONS


def run_parse(filename, year_from_user):
    wb = opx.load_workbook(filename)
    title_sheet = wb["Титул"]
    plan_sheet = wb["План"]
    year_from_user = int(year_from_user)

    entry_year = find_entry_year(sheet=title_sheet)
    group_name, graduate_year = get_group_name(filename=filename, entry_year=entry_year)

    if year_from_user > graduate_year:
        return False

    education_form = find_education_form(sheet=title_sheet)

    first_semester, second_semester = find_semester_boundaries(sheet=plan_sheet, today_year=year_from_user,
                                                               entry_year=int(entry_year))

    disciplines_col = find_disciplines_column(sheet=plan_sheet)

    study_load = parse_study_load(sheet=plan_sheet, first_semester=first_semester, second_semester=second_semester,
                                  disciplines_col=disciplines_col)

    control_boundaries = find_form_control_boundaries(sheet=plan_sheet)
    first_semester_control = parse_form_control(sheet=plan_sheet, form_control_boundaries=control_boundaries,
                                                semester=first_semester[5])
    second_semester_control = parse_form_control(sheet=plan_sheet, form_control_boundaries=control_boundaries,
                                                 semester=second_semester[5])
    result = format_to_converter(study_load=study_load, group_name=group_name, education_form=education_form,
                                 first_semester_control=first_semester_control,
                                 second_semester_control=second_semester_control)

    print("Имя файла:", filename)
    print("Год поступления: ", entry_year)
    print("Название группы: ", group_name)
    print("Форма обучения: ", education_form)
    print("Границы первого семестра: ", first_semester)
    print("Границы второго семестра: ", second_semester)
    
    path = get_filepath((int(year_from_user)))
    tf = opx.load_workbook(path)
    tf_ws = tf['Тарификация']
    with pd.ExcelWriter(path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        result.to_excel(writer, sheet_name="Тарификация", header=False, index=False, startrow=tf_ws.max_row)
    return True


def finilaze_converted_file(year_from_user):
    path = get_filepath(year_from_user)
    wb = opx.load_workbook(path)
    ws = wb.active
    last_row = ws.max_row

    for row in range(13, last_row + 1):
        total_aud = f"=SUM(M{row},N{row},R{row})"
        total = f"=SUM(S{row},O{row},P{row})"
        additional = f"=T{row}/K{row}"
        rounded_additional = f"=ROUND(V{row},0)"
        ws[f"S{row}"] = total_aud
        ws[f"T{row}"] = total
        ws[f"V{row}"] = additional
        ws[f"W{row}"] = rounded_additional

    subtotal_columns = ['M', 'N', 'O', 'P', 'Q', 'R', 'T']
    for column in subtotal_columns:
        ws[f'{column}{last_row + 2}'] = f'=SUBTOTAL(9,{column}13:{column}{last_row})'
    ws[f'S{last_row + 2}'] = f'=SUM(M{last_row + 2}:N{last_row + 2}) + R{last_row + 2}'
    ws[f'S{last_row + 3}'] = f'=S{last_row + 2}/720'
    ws[f'T{last_row + 3}'] = f'=T{last_row + 2}/720'

    ws[f'V{last_row + 2}'] = f'=SUBTOTAL(9,V13:V{last_row})'
    ws[f'W{last_row + 2}'] = f'=SUBTOTAL(9,W13:W{last_row})'
    ws[f'V{last_row + 3}'] = f'=V{last_row + 2}/2'
    ws[f'W{last_row + 3}'] = f'=W{last_row + 2}/2'
    wb.save(path)


def check_filenames(filesnames: Any) -> List[str]:
    """Проверяет список файлов с которыми мы будем работать на корректность названия

    :param filesnames: Список файлов которые выбрал пользователь
    :type filesnames: List
    :returns: Список файлов прошедших валидацию
    :rtype: list
    """
    valid_files = []
    invalid_files = []
    for filename in filesnames:
        if filename_validator(filename):
            valid_files.append(filename)
            print(f"[+]: {filename} валидно")
        else:
            invalid_files.append(filename)
            print(f"[ERR]: {filename} не валидно")
    return valid_files


def filename_validator(filename: str) -> bool:
    """
    # Код.специальности_XXX_стандарт_количество.курсов_XXX_класс_год.набора_специальность
    # regex: d{2}.d{2}.d{2}_d{2}_d{2}_(d{3,4}/d{4})-2843_(d{2})-(d{4})

    :param filename: Имя файла которое будем валидировать
    :type filename: string
    :returns: Значение результата проверки
    :rtype: bool
    """
    if re.search(r"\d{2}\.\d{2}\.\d{2}[_-]\d{2}[_-]\d{2}[_-]\d{2,4}[_-]2843[_-]\d{2}[_-]\d{4}[_-][а-яА-Я]{1,3}",
                 filename):
        return True
    return False


def get_group_name(filename: str, entry_year: int) -> List[str]:
    """Из имени файла извлекается код специальности, база класса,
    и год поступления передается из результата другой функции. В результате формируется имя группы:
    год-сокращение_специальности-класса"""

    file_name = filename.split("/")[-1]
    file_data_list = file_name.replace("_", "-").rstrip(".").split("-")[:-1]
    file_data_list.pop(1)  # Удаление ненужной части
    file_data_list.remove("2843")  # Удаления кода организации "2843"
    specialization_key = file_data_list[0]
    graduate_year = int(entry_year) + int(len(file_data_list[2]))
    group_classbase = file_data_list[3]
    if group_classbase == "09":
        group_classbase = 9
    group_codename = SPECIALIZATIONS[specialization_key][0]
    group_name = f"{entry_year[-2:]}-{group_codename}-{group_classbase}"
    return [group_name, graduate_year]


def find_education_form(sheet: Worksheet) -> Optional[str]:
    forms_dict = {"Очная": "ОФО",
                  "Очно-заочная": "ОЗФО",
                  "Заочная": "ЗФО"}
    pattern = re.compile(r'(Очная|очно-заочная|заочная)')
    for row in sheet.iter_rows(min_row=30, max_row=60, values_only=True):
        for cell in row:
            if cell and pattern.search(str(cell)):
                education_form = pattern.search(str(cell)).group(1)
                if education_form in forms_dict:
                    education_form = forms_dict[education_form]
                    return education_form
    return None


def find_entry_year(sheet: Worksheet) -> Optional[int]:
    pattern = re.compile(r"^\d{4}$")
    for row in sheet.iter_rows(min_row=30, max_row=60, values_only=True):
        year_cell = next((cell for cell in row if cell and re.match(pattern, str(cell))), None)
        if year_cell:
            return year_cell
    return None


def find_semester_boundaries(sheet: Worksheet, today_year: int, entry_year: int):
    course_number = today_year - entry_year + 1
    course_boundaries = []
    semester_data = []
    semester_number = ""
    semester_length = ""
    pattern = re.compile(r"^[Кк]урс\s+" + str(course_number))

    for row in sheet.iter_rows(max_row=5):
        for cell in row:
            if isinstance(cell.value, str) and pattern.search(cell.value):
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        course_boundaries.extend([merged_range.min_col, merged_range.max_col, merged_range.min_row,
                                                  merged_range.max_row])
        break

    for row in sheet.iter_rows(min_col=course_boundaries[0], max_col=course_boundaries[1], min_row=course_boundaries[2],
                               max_row=course_boundaries[3] + 1):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("Семестр"):
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        try:
                            match = re.search(r'Семестр\s+(\d+)', cell.value)
                            semester_number = int(match.group(1))
                        except Exception:
                            print("На листе не была найдена ячейка со значением 'Семестр'")
                        try:
                            match = re.search(r'(?<=\[)\d+', cell.value)
                            semester_length = int(match.group(0))
                        except Exception:
                            print(
                                f"На листе не было найдено значение длинны для семестра '{cell.value}' "
                                f"в ячейке: {cell.coordinate}")
                        semester_data.append([merged_range.min_col, merged_range.max_col, merged_range.min_row,
                                              merged_range.max_row, course_number, semester_number, semester_length])

    return semester_data


def find_disciplines_column(sheet: Worksheet) -> int:
    pattern = re.compile(r'Наименование$')
    for row in sheet.iter_rows(min_row=1, max_row=10, max_col=5):
        for cell in row:
            if isinstance(cell.value, str) and pattern.search(cell.value):
                return cell.column


def find_form_control_boundaries(sheet: Worksheet) -> List[int]:
    pattern = re.compile(r'(Форма|контроля)', flags=re.IGNORECASE)
    control_boundaries = []

    for row in sheet.iter_rows(max_row=5):
        for cell in row:
            if isinstance(cell.value, str) and pattern.search(cell.value):
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        control_boundaries.append([merged_range.min_col, merged_range.max_col, merged_range.min_row,
                                                   merged_range.max_row])
    return control_boundaries[0]


def parse_form_control(sheet: Worksheet, form_control_boundaries: List[int], semester: int) -> DataFrame:
    data = []
    # (columns=["Шифр дисциплины", "Форма контроля"])
    disciplines_pattern = re.compile(r'[а-яА-Я]+\.\d{1,3}')

    for row in sheet.iter_rows(max_col=form_control_boundaries[1]):
        if row[0].value and row[1].value:
            if not row[2].font.bold:
                match = disciplines_pattern.search(row[1].value)
                if match:
                    data.append([row[1].value, *[cell.value for cell in row[form_control_boundaries[0] - 1:
                                                                            form_control_boundaries[1]]]])
    if len(data[0]) == 7:
        result = pd.DataFrame(data, columns=["Шифр дисциплины", "экз", "зач", "д/зач", "кп", "кр", "Оценка"])
    elif len(data[0]) == 6:
        result = pd.DataFrame(data, columns=["Шифр дисциплины", "экз", "зач", "д/зач", "кр", "Оценка"])
    else:
        result = pd.DataFrame(data, columns=["Шифр дисциплины", "экз", "зач", "д/зач", "кр"])

    result = result.set_index('Шифр дисциплины').stack().astype(str).apply(lambda x: list(x)).explode()
    result = result.reset_index().rename(columns={'level_1': 'Форма контроля', 0: 'Семестр'})
    result = result[result['Семестр'] != 'nan'].reset_index(drop=True)
    result = result.query(f"Семестр == '{semester}'")
    return result


# def parse_disciplines(sheet: Worksheet) -> DataFrame:
#     data = pd.DataFrame(columns=["Шифр дисциплины", "Наименование дисциплины"])
#     pattern = re.compile(r"[а-яА-Я]+\.\d{1,3}")  # https://regex101.com/r/Htwvzy/1
#
#     def is_discipline(cell):
#         return cell.value and isinstance(cell.value, str) and pattern.search(cell.value)
#
#     for row in sheet.iter_rows():
#         discipline_cell = next((cell for cell in row if is_discipline(cell)), None)
#         if not discipline_cell:
#             continue
#         data.loc[len(data)] = [discipline_cell.value, ""]
#         last_idx = len(data) - 1
#         next_cell_index = discipline_cell.column + 1
#         if next_cell_index <= row[-1].column:
#             data.at[last_idx, "Наименование дисциплины"] = row[next_cell_index - 1].value
#
#     return data


def parse_study_load(
        sheet: Worksheet, first_semester: List[int], second_semester: List[int],
        disciplines_col: int) -> List[DataFrame]:
    disciplines_data = []
    first_semester_data = []
    second_semester_data = []
    disciplines_pattern = re.compile(r'[а-яА-Я]+\.\d{1,3}')
    # Парсинг 1 семестра
    for row in sheet.iter_rows(min_col=disciplines_col - 1, max_col=first_semester[1], values_only=False):
        if row[0].value and row[1].value:
            if not row[1].font.bold:
                match = disciplines_pattern.search(row[0].value)
                if match:
                    disciplines_data.append([row[0].value, row[1].value])
                    # Черная магия
                    first_semester_data.append(
                        [row[0].value, first_semester[4], first_semester[5], first_semester[6],
                         *[cell.value for cell in row[first_semester[0] - 2:first_semester[1] + 1]]]
                    )

    # Парсинг 2 семестра
    for row in sheet.iter_rows(min_col=disciplines_col - 1, max_col=second_semester[1], values_only=False):
        if row[0].value and row[1].value:
            if not row[1].font.bold:
                match = disciplines_pattern.search(row[0].value)
                if match:
                    # Черная магия
                    second_semester_data.append(
                        [row[0].value, second_semester[4], second_semester[5], second_semester[6],
                         *[cell.value for cell in row[second_semester[0] - 2:second_semester[1] + 1]]]
                    )

    disciplines_df = pd.DataFrame(disciplines_data,
                                  columns=["Шифр дисциплины", "Наименование дисциплины"])
    # int_columns = ["Курс", "Семестр", "Число учебных недель", "Всего часов", "Всего ауд. часов", "Лекции",
    #                "Практические занятия", "КРП", "ИП", "Экзамены"]
    int_columns = ["Курс", "Семестр", "Число учебных недель", "Всего часов", "Всего ауд. часов",
                   "Лекции", "Лаб", "Практические занятия", "КРП", "ИП", "Консультации", "Экзамены"]
    if len(first_semester_data[0]) == 13:
        columns = ["Шифр дисциплины", "Курс", "Семестр",
                   "Число учебных недель", "Всего часов", "Всего ауд. часов", "Лекции",
                   "Практические занятия", "КРП",
                   "ИП", "Консультации", "Сам.работа (по актуализ.ФГОС)", "Экзамены"]
    elif len(first_semester_data[0]) == 12:
        columns = ["Шифр дисциплины", "Курс", "Семестр",
                   "Число учебных недель", "Всего часов", "Всего ауд. часов", "Лекции",
                   "Практические занятия", "КРП",
                   "ИП", "Сам.работа (по актуализ.ФГОС)", "Экзамены"]
    elif len(first_semester_data[0]) == 11:
        columns = ["Шифр дисциплины", "Курс", "Семестр",
                   "Число учебных недель", "Всего часов", "Всего ауд. часов", "Лекции",
                   "Практические занятия", "КРП",
                   "ИП", "Сам.работа (по актуализ.ФГОС)"]
    else:
        columns = ["Шифр дисциплины", "Курс", "Семестр",
                   "Число учебных недель", "Всего часов", "Всего ауд. часов", "Лекции", "Лаб",
                   "Практические занятия", "КРП",
                   "ИП", "Консультации", "Сам.работа (по актуализ.ФГОС)", "Экзамены"]

    first_semester_df = pd.DataFrame(first_semester_data, columns=columns)
    second_semester_df = pd.DataFrame(second_semester_data, columns=columns)

    def convert_to_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    for col in int_columns:
        if col in first_semester_df.columns:
            first_semester_df[col] = first_semester_df[col].apply(convert_to_int)
        if col in second_semester_df.columns:
            second_semester_df[col] = second_semester_df[col].apply(convert_to_int)

    return [disciplines_df, first_semester_df, second_semester_df]


# TODO доделать вторую половину таблицы, понять откуда парсить "Продолжительность практики (нед)",
def format_to_converter(study_load: List[DataFrame], group_name: str, education_form: str,
                        first_semester_control, second_semester_control):
    disciplines_df, first_semester_df, second_semester_df = study_load
    # print(disciplines_df.head())
    # print(first_semester_df.head())
    # print(second_semester_df.head())
    cols = pd.DataFrame()
    result = pd.concat([cols, disciplines_df[["Наименование дисциплины", "Шифр дисциплины"]]], axis=1)
    result["Форма обучения"] = education_form
    result["Номер группы"] = group_name
    result["Ф.И.О. преподавателя"] = 0
    result['Продолжительность практики (нед)'] = 0

    if first_semester_df.shape[1] == 13 or first_semester_df.shape[1] == 14:
        fsem_mask = first_semester_df["Всего часов"].notna() | first_semester_df["Всего ауд. часов"].notna() | \
                    first_semester_df['Лекции'].notna() | first_semester_df['Практические занятия'].notna() | \
                    first_semester_df['КРП'].notna() | first_semester_df['ИП'].notna() | \
                    first_semester_df['Консультации'].notna() | \
                    first_semester_df['Сам.работа (по актуализ.ФГОС)'].notna() | \
                    first_semester_df['Экзамены'].notna()

        ssem_mask = second_semester_df["Всего часов"].notna() | second_semester_df["Всего ауд. часов"].notna() | \
                    second_semester_df['Лекции'].notna() | second_semester_df['Практические занятия'].notna() | \
                    second_semester_df['КРП'].notna() | second_semester_df['ИП'].notna() | \
                    second_semester_df['Консультации'].notna() | \
                    second_semester_df['Сам.работа (по актуализ.ФГОС)'].notna() | \
                    second_semester_df['Экзамены'].notna()
    elif first_semester_df.shape[1] == 11:
        result["Консультации"] = 0
        result["Экзамены"] = 0
        fsem_mask = first_semester_df["Всего часов"].notna() | first_semester_df["Всего ауд. часов"].notna() | \
                    first_semester_df['Лекции'].notna() | first_semester_df['Практические занятия'].notna() | \
                    first_semester_df['КРП'].notna() | first_semester_df['ИП'].notna() | \
                    first_semester_df['Сам.работа (по актуализ.ФГОС)'].notna()

        ssem_mask = second_semester_df["Всего часов"].notna() | second_semester_df["Всего ауд. часов"].notna() | \
                    second_semester_df['Лекции'].notna() | second_semester_df['Практические занятия'].notna() | \
                    second_semester_df['КРП'].notna() | second_semester_df['ИП'].notna() | \
                    second_semester_df['Сам.работа (по актуализ.ФГОС)'].notna()

    else:
        result["Консультации"] = 0
        fsem_mask = first_semester_df["Всего часов"].notna() | first_semester_df["Всего ауд. часов"].notna() | \
                    first_semester_df['Лекции'].notna() | first_semester_df['Практические занятия'].notna() | \
                    first_semester_df['КРП'].notna() | first_semester_df['ИП'].notna() | \
                    first_semester_df['Сам.работа (по актуализ.ФГОС)'].notna() | \
                    first_semester_df['Экзамены'].notna()

        ssem_mask = second_semester_df["Всего часов"].notna() | second_semester_df["Всего ауд. часов"].notna() | \
                    second_semester_df['Лекции'].notna() | second_semester_df['Практические занятия'].notna() | \
                    second_semester_df['КРП'].notna() | second_semester_df['ИП'].notna() | \
                    second_semester_df['Сам.работа (по актуализ.ФГОС)'].notna() | \
                    second_semester_df['Экзамены'].notna()

    fsem = pd.merge(result[fsem_mask], first_semester_df, on='Шифр дисциплины')
    ssem = pd.merge(result[ssem_mask], second_semester_df, on='Шифр дисциплины')
    f_control = pd.merge(fsem, first_semester_control[["Шифр дисциплины", "Форма контроля"]], on="Шифр дисциплины",
                         how="left")
    s_control = pd.merge(ssem, second_semester_control[["Шифр дисциплины", "Форма контроля"]], on="Шифр дисциплины",
                         how="left")
    result = pd.concat([f_control, s_control], ignore_index=True)

    # print(result.to_string())
    for i in range(6, 9):
        result.insert(loc=i, column=COLUMNS[i], value="")

    result["КРП"] = result["КРП"].fillna(0).astype(int)
    result["ИП"] = result["ИП"].fillna(0).astype(int)
    _ = result["КРП"] + result["ИП"]
    result = result.assign(КРП=_).rename(columns={"КРП": COLUMNS[17]})
    result = result.drop(["ИП"], axis=1)
    result['Сам.работа (по актуализ.ФГОС)'] = " "
    result = result[["Наименование дисциплины", "Шифр дисциплины", "Форма обучения", "Номер группы",
                     "Курс", "Семестр",
                     "Студентов", "Потоков", "Групп",
                     "Продолжительность практики (нед)", "Число учебных недель",
                     "Форма контроля", "Лекции", "Практические занятия", "Консультации",
                     "Экзамены", "Сам.работа (по актуализ.ФГОС)", "Курсовые работы / Индивидуальный проект",
                     "Всего ауд. часов", "Всего часов", "Ф.И.О. преподавателя"]]
    result = result.replace(0, None)
    order = list(result['Шифр дисциплины'].str.split('.').str.get(0).unique())
    print("order: ", order)
    if len(order) > 1:
        cat = pd.Categorical(result['Шифр дисциплины'].str.split('.').str[0],
                             categories=order)
        result2 = result.groupby(cat).apply(
            lambda x: x
            .sort_values('Шифр дисциплины', key=lambda x: x.str.split('.').str[1]))\
            .reset_index(drop=True)
        # print("result\n", result)

    else:
        result2 = result.sort_values(by=['Шифр дисциплины'],)
    print("result2\n", result2)
    return result2
